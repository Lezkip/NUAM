import json
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import ProtectedError
from django.db import transaction
from .models import Emisor, FactorTributario, Calificacion, HistorialAuditoria
from .forms import EmisorForm, FactorForm, CalificacionForm, CargaMasivaForm

# ==========================================
# 0. DASHBOARD
# ==========================================

@login_required
def dashboard(request):
    return render(request, 'calificaciones/dashboard.html')

# ==========================================
# 1. DECORADORES Y AYUDAS
# ==========================================

def require_group(user, group_name):
    return user.is_authenticated and user.groups.filter(name=group_name).exists()

def analyst_required(view_func):
    def _wrapped(request, *args, **kwargs):
        # Si es superusuario o tiene is_staff (Analista), pasa.
        if request.user.is_superuser or request.user.is_staff:
            return view_func(request, *args, **kwargs)
        else:
            return render(
                request,
                "calificaciones/forbidden.html",
                {
                    "title": "Acceso restringido",
                    "reason": "No tienes permisos de Analista para realizar esta acción.",
                    "action": "Contacta a tu administrador."
                },
                status=403,
            )
    return _wrapped

# ==========================================
# 2. GESTIÓN DE EMISORES
# ==========================================

@login_required
def lista_emisores(request):
    # Filtra solo los activos (baja lógica)
    emisores = Emisor.objects.filter(activo=True)
    return render(request, 'calificaciones/lista_emisores.html', {'emisores': emisores})

@login_required
def detalle_emisor(request, id):
    try:
        emisor = Emisor.objects.values("id", "rut", "nombre", "direccion").get(id=id)
        return JsonResponse(emisor, safe=False)
    except Emisor.DoesNotExist:
        return JsonResponse({"error": "Emisor no encontrado"}, status=404)

@login_required
@analyst_required
def emisor_create(request):
    if request.method == "POST":
        form = EmisorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Emisor creado correctamente.")
            return redirect("lista_emisores")
    else:
        form = EmisorForm()
    return render(request, "calificaciones/form_generico.html", {"form": form, "title": "Nuevo Emisor"})

@login_required
@analyst_required
def emisor_update(request, pk):
    emisor = get_object_or_404(Emisor, pk=pk)
    if request.method == "POST":
        form = EmisorForm(request.POST, instance=emisor)
        if form.is_valid():
            form.save()
            messages.success(request, "Emisor actualizado.")
            return redirect("lista_emisores")
    else:
        form = EmisorForm(instance=emisor)
    return render(request, "calificaciones/form_generico.html", {"form": form, "title": "Editar Emisor"})

@login_required
@analyst_required
def emisor_delete(request, pk):
    emisor = get_object_or_404(Emisor, pk=pk)
    # Baja lógica si tiene calificaciones
    tiene_calif = Calificacion.objects.filter(emisor=emisor).exists()
    
    if request.method == "POST":
        if tiene_calif:
            emisor.activo = False
            emisor.save()
            messages.warning(request, "El emisor tiene historial. Se realizó una baja lógica.")
        else:
            emisor.delete()
            messages.success(request, "Emisor eliminado permanentemente.")
        return redirect("lista_emisores")
    
    return render(request, "calificaciones/confirm_delete.html", {"obj": emisor, "logical": tiene_calif})


# ==========================================
# 3. GESTIÓN DE FACTORES
# ==========================================

@login_required
def lista_factores(request):
    factores = FactorTributario.objects.all()
    return render(request, 'calificaciones/lista_factores.html', {'factores': factores})

@login_required
def detalle_factor(request, id):
    try:
        factor = FactorTributario.objects.values("id", "codigo", "descripcion", "vigente").get(id=id)
        return JsonResponse(factor, safe=False)
    except FactorTributario.DoesNotExist:
        return JsonResponse({"error": "Factor no encontrado"}, status=404)

@login_required
@analyst_required
def factor_create(request):
    if request.method == "POST":
        form = FactorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Factor creado correctamente.")
            return redirect("lista_factores")
    else:
        form = FactorForm()
    return render(request, "calificaciones/form_generico.html", {"form": form, "title": "Nuevo Factor"})

@login_required
@analyst_required
def factor_update(request, pk):
    factor = get_object_or_404(FactorTributario, pk=pk)
    codigo_anterior = factor.codigo
    descripcion_anterior = factor.descripcion
    if request.method == "POST":
        form = FactorForm(request.POST, instance=factor)
        if form.is_valid():
            form.save()

            # Registrar auditoría de cambio de código o descripción de factor
            codigo_nuevo = factor.codigo
            descripcion_nueva = factor.descripcion
            if codigo_anterior != codigo_nuevo or descripcion_anterior != descripcion_nueva:
                emisor_ids = Calificacion.objects.filter(factor=factor).values_list('emisor_id', flat=True).distinct()
                
                # Construir mensaje de cambios
                cambios = []
                if codigo_anterior != codigo_nuevo:
                    cambios.append(f"Código: {codigo_anterior} → {codigo_nuevo}")
                if descripcion_anterior != descripcion_nueva:
                    cambios.append(f"Descripción: {descripcion_anterior} → {descripcion_nueva}")
                mensaje_cambios = " | ".join(cambios)
                
                registros = [
                    HistorialAuditoria(
                        usuario=request.user,
                        emisor_id=emisor_id,
                        factor_anterior=codigo_anterior,
                        factor_nuevo=codigo_nuevo,
                        comentario_anterior=descripcion_anterior,
                        comentario_nuevo=mensaje_cambios
                    )
                    for emisor_id in emisor_ids
                ]
                if registros:
                    HistorialAuditoria.objects.bulk_create(registros)

            messages.success(request, "Factor actualizado.")
            return redirect("lista_factores")
    else:
        form = FactorForm(instance=factor)
    return render(request, "calificaciones/form_generico.html", {"form": form, "title": "Editar Factor"})

@login_required
@analyst_required
def factor_delete(request, pk):
    factor = get_object_or_404(FactorTributario, pk=pk)
    # Validación: No eliminar si está en uso
    if Calificacion.objects.filter(factor=factor).exists():
        messages.error(request, "No se puede eliminar: El factor está asociado a calificaciones.")
        return redirect("lista_factores")

    if request.method == "POST":
        factor.delete()
        messages.success(request, "Factor eliminado.")
        return redirect("lista_factores")
        
    return render(request, "calificaciones/confirm_delete.html", {"obj": factor, "logical": False})


# ==========================================
# 4. GESTIÓN DE CALIFICACIONES
# ==========================================

@login_required
def lista_calificaciones(request):
    calificaciones = Calificacion.objects.select_related('emisor', 'factor', 'usuario').all()
    return render(request, 'calificaciones/lista_calificaciones.html', {'calificaciones': calificaciones})

@login_required
def detalle_calificacion(request, id):
    
    try:
        c = Calificacion.objects.select_related('emisor', 'factor').get(id=id)
        data = {
            "id": c.id,
            "emisor": c.emisor.nombre,
            "rut": c.emisor.rut,
            "factor": f"{c.factor.codigo} - {c.factor.descripcion}",
            "fecha": c.fecha_asignacion.strftime("%Y-%m-%d %H:%M"),
            "usuario": c.usuario.username if c.usuario else "Sistema",
            "comentario": c.comentario or "Sin comentarios"
        }
        return JsonResponse(data)
    except Calificacion.DoesNotExist:
        return JsonResponse({"error": "Calificación no encontrada"}, status=404)

@login_required
@analyst_required
def calificacion_create(request):
    if request.method == "POST":
        form = CalificacionForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.usuario = request.user
            obj.save()
            
            # Registrar Auditoría de Creación
            HistorialAuditoria.objects.create(
                usuario=request.user,
                emisor=obj.emisor,
                factor_anterior="N/A",
                factor_nuevo=obj.factor.codigo
            )
            
            messages.success(request, "Calificación creada.")
            return redirect("calificacion_list")
    else:
        form = CalificacionForm()
    return render(request, "calificaciones/form_generico.html", {"form": form, "title": "Nueva Calificación"})

@login_required
@analyst_required
def calificacion_update(request, pk):
    obj = get_object_or_404(Calificacion, pk=pk)
    factor_anterior = obj.factor.codigo
    comentario_anterior = obj.comentario or ""
    
    if request.method == "POST":
        form = CalificacionForm(request.POST, instance=obj)
        if form.is_valid():
            nueva_calif = form.save(commit=False)
            nueva_calif.usuario = request.user
            nueva_calif.save()
            
            # Registrar Auditoría si cambió el factor o el comentario
            if factor_anterior != nueva_calif.factor.codigo or comentario_anterior != (nueva_calif.comentario or ""):
                HistorialAuditoria.objects.create(
                    usuario=request.user,
                    emisor=nueva_calif.emisor,
                    factor_anterior=factor_anterior,
                    factor_nuevo=nueva_calif.factor.codigo,
                    comentario_anterior=comentario_anterior,
                    comentario_nuevo=nueva_calif.comentario or ""
                )
            
            messages.success(request, "Calificación actualizada.")
            return redirect("calificacion_list")
    else:
        form = CalificacionForm(instance=obj)
    return render(request, "calificaciones/form_generico.html", {"form": form, "title": "Editar Calificación"})

@login_required
@analyst_required
def calificacion_delete(request, pk):
    obj = get_object_or_404(Calificacion, pk=pk)
    if request.method == "POST":
        # Auditoría de eliminación
        HistorialAuditoria.objects.create(
            usuario=request.user,
            emisor=obj.emisor,
            factor_anterior=obj.factor.codigo,
            factor_nuevo="ELIMINADO"
        )
        obj.delete()
        messages.success(request, "Calificación eliminada.")
        return redirect("calificacion_list")
    return render(request, "calificaciones/confirm_delete.html", {"obj": obj, "logical": False})


# ==========================================
# 5. CARGA MASIVA (API)
# ==========================================

@login_required
def api_carga_masiva(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        
        if not archivo.name.endswith('.xlsx'):
            return JsonResponse({'error': 'Formato inválido. Use .xlsx'}, status=400)

        try:
            # 1. Cargar datos en memoria (Cache)
            emisores_map = {e.rut: e for e in Emisor.objects.all()}
            factores_map = {f.codigo: f for f in FactorTributario.objects.all()}
            
            # 2. Leer Excel optimizado
            wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            sheet = wb.active
            
            nuevas_calificaciones = []
            errores = []
            filas_procesadas = 0

            # 3. Iterar filas
            # values_only=True devuelve una tupla con los valores de la fila
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                
                # --- CORRECCIÓN DEL ERROR INDEX OUT OF RANGE ---
                # Si la fila está vacía o tiene menos de 2 columnas, la saltamos
                if not row or len(row) < 2:
                    continue
                
                rut = row[0]
                codigo_factor = row[1]
                
                # Si las celdas están vacías (None), saltar
                if not rut or not codigo_factor:
                    continue

                emisor = emisores_map.get(rut)
                factor = factores_map.get(codigo_factor)

                if not emisor:
                    # Opcional: limitar errores para no llenar la memoria si todo falla
                    if len(errores) < 50: errores.append(f"Fila {i}: RUT {rut} no encontrado")
                    continue
                if not factor:
                    if len(errores) < 50: errores.append(f"Fila {i}: Factor {codigo_factor} no encontrado")
                    continue

                nuevas_calificaciones.append(
                    Calificacion(emisor=emisor, factor=factor, usuario=request.user)
                )
                filas_procesadas += 1

                # Guardar en lotes de 5,000 para liberar memoria y evitar que se pegue
                if len(nuevas_calificaciones) >= 5000:
                    with transaction.atomic():
                        Calificacion.objects.bulk_create(nuevas_calificaciones, ignore_conflicts=True)
                    nuevas_calificaciones = []
                    print(f"Procesadas {filas_procesadas} filas...") # Log en consola para ver avance

            # Guardar el último grupo restante
            if nuevas_calificaciones:
                with transaction.atomic():
                    Calificacion.objects.bulk_create(nuevas_calificaciones, ignore_conflicts=True)

            # 4. Auditoría (Solo 1 registro)
            if filas_procesadas > 0:
                HistorialAuditoria.objects.create(
                    usuario=request.user,
                    emisor=emisores_map.get(list(emisores_map.keys())[0]) if emisores_map else None,
                    factor_nuevo="CARGA_MASIVA",
                    factor_anterior=f"{filas_procesadas} registros cargados exitosamente."
                )

            wb.close()
            
            mensaje = f"Éxito: {filas_procesadas} registros cargados."
            if errores:
                mensaje += f" (Se omitieron {len(errores)} filas con errores)."
                
            return JsonResponse({'mensaje': mensaje, 'errores': errores[:10]})

        except Exception as e:
            print(f"Error crítico: {e}") # Ver error en terminal
            return JsonResponse({'error': f"Error en el servidor: {str(e)}"}, status=500)

    return JsonResponse({'error': 'Petición inválida'}, status=400)


# ==========================================
# 5. AUDITORÍA
# ==========================================

@login_required
def lista_auditoria(request):
    # Solo Analista (is_staff) puede ver el historial
    if not (request.user.is_superuser or request.user.is_staff):
        return render(
            request,
            "calificaciones/forbidden.html",
            {
                "title": "Acceso restringido",
                "reason": "Solo usuarios Analista pueden ver el historial de cambios.",
                "action": "Contacta a tu administrador si necesitas acceso."
            },
            status=403,
        )
    
    historial = HistorialAuditoria.objects.select_related('usuario', 'emisor').all()
    return render(request, 'calificaciones/lista_auditoria.html', {'historial': historial})


# ==========================================
# 5. CARGA MASIVA DE CALIFICACIONES
# ==========================================

@login_required
@analyst_required
def carga_masiva_calificaciones(request):
    """Vista para cargar calificaciones desde archivo Excel o CSV"""
    if request.method == "POST":
        form = CargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            try:
                # Determinar el tipo de archivo
                nombre_archivo = archivo.name.lower()
                if nombre_archivo.endswith('.csv'):
                    resultado = procesar_csv_calificaciones(archivo, request.user)
                else:  # Excel
                    resultado = procesar_excel_calificaciones(archivo, request.user)
                
                messages.success(request, resultado['mensaje'])
                
                # Mostrar detalles de la carga
                if resultado['errores']:
                    for error in resultado['errores']:
                        messages.warning(request, f"⚠ Fila {error['fila']}: {error['mensaje']}")
                
                return redirect('calificacion_list')
            except Exception as e:
                messages.error(request, f"Error al procesar archivo: {str(e)}")
        else:
            # Mostrar errores del formulario (incluidos los de validación de archivo)
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"❌ {error}")
    else:
        form = CargaMasivaForm()
    
    return render(request, 'calificaciones/carga_masiva.html', {'form': form})


def _maps_emisores_factores():
    """Carga emisores y factores en RAM para lookups O(1)."""
    emisores_map = {e.rut: e for e in Emisor.objects.all()}
    factores_map = {f.codigo: f for f in FactorTributario.objects.all()}
    return emisores_map, factores_map


def _bulk_insert_calificaciones(pendientes, creadas):
    """Inserta en lote con ignore_conflicts para saltar duplicados."""
    if not pendientes:
        return creadas
    with transaction.atomic():
        Calificacion.objects.bulk_create(pendientes, ignore_conflicts=True)
    return creadas + len(pendientes)


def procesar_excel_calificaciones(archivo, usuario, chunk_size=2000):
    """Procesa Excel con lookups en memoria y bulk_create por lotes."""
    workbook = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    worksheet = workbook.active

    emisores_map, factores_map = _maps_emisores_factores()
    existentes = set(Calificacion.objects.values_list('emisor_id', 'factor_id'))

    pendientes = []
    creadas = 0
    errores = []
    errores_totales = 0

    for fila_num, fila in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not fila or not fila[0] or not fila[1]:
                continue

            rut = str(fila[0]).strip()
            codigo_factor = str(fila[1]).strip()
            comentario = fila[2] if len(fila) > 2 else ""

            emisor = emisores_map.get(rut)
            if not emisor:
                errores_totales += 1
                if len(errores) < 100:
                    errores.append({'fila': fila_num, 'mensaje': f"Emisor con RUT {rut} no encontrado"})
                continue

            factor = factores_map.get(codigo_factor)
            if not factor:
                errores_totales += 1
                if len(errores) < 100:
                    errores.append({'fila': fila_num, 'mensaje': f"Factor {codigo_factor} no encontrado"})
                continue

            clave = (emisor.id, factor.id)
            if clave in existentes:
                continue

            pendientes.append(
                Calificacion(
                    emisor=emisor,
                    factor=factor,
                    usuario=usuario,
                    comentario=comentario or ""
                )
            )
            existentes.add(clave)

            if len(pendientes) >= chunk_size:
                creadas = _bulk_insert_calificaciones(pendientes, creadas)
                pendientes = []

        except Exception as e:
            errores_totales += 1
            if len(errores) < 100:
                errores.append({'fila': fila_num, 'mensaje': str(e)})

    # Insertar remanentes
    creadas = _bulk_insert_calificaciones(pendientes, creadas)
    workbook.close()

    mensaje = f"✓ {creadas} calificaciones cargadas exitosamente"
    if errores_totales:
        mensaje += f" ({errores_totales} errores)"

    return {
        'creadas': creadas,
        'errores': errores,
        'mensaje': mensaje
    }


def procesar_csv_calificaciones(archivo, usuario, chunk_size=2000):
    """Procesa CSV con lookups en memoria y bulk_create por lotes."""
    import csv
    import io

    emisores_map, factores_map = _maps_emisores_factores()
    existentes = set(Calificacion.objects.values_list('emisor_id', 'factor_id'))

    creadas = 0
    errores = []
    errores_totales = 0
    pendientes = []

    contenido = archivo.read().decode('utf-8')
    reader = csv.DictReader(io.StringIO(contenido))

    for fila_num, fila in enumerate(reader, start=2):
        try:
            rut = (fila.get('RUT', '') or '').strip()
            codigo_factor = (fila.get('Código Factor', '') or '').strip()
            comentario = (fila.get('Comentario', '') or '').strip()

            if not rut or not codigo_factor:
                continue

            emisor = emisores_map.get(rut)
            if not emisor:
                errores_totales += 1
                if len(errores) < 100:
                    errores.append({'fila': fila_num, 'mensaje': f"Emisor con RUT {rut} no encontrado"})
                continue

            factor = factores_map.get(codigo_factor)
            if not factor:
                errores_totales += 1
                if len(errores) < 100:
                    errores.append({'fila': fila_num, 'mensaje': f"Factor {codigo_factor} no encontrado"})
                continue

            clave = (emisor.id, factor.id)
            if clave in existentes:
                continue

            pendientes.append(
                Calificacion(
                    emisor=emisor,
                    factor=factor,
                    usuario=usuario,
                    comentario=comentario or ""
                )
            )
            existentes.add(clave)

            if len(pendientes) >= chunk_size:
                creadas = _bulk_insert_calificaciones(pendientes, creadas)
                pendientes = []

        except Exception as e:
            errores_totales += 1
            if len(errores) < 100:
                errores.append({'fila': fila_num, 'mensaje': str(e)})

    creadas = _bulk_insert_calificaciones(pendientes, creadas)

    mensaje = f"✓ {creadas} calificaciones cargadas exitosamente"
    if errores_totales:
        mensaje += f" ({errores_totales} errores)"

    return {
        'creadas': creadas,
        'errores': errores,
        'mensaje': mensaje
    }


# ==========================================
# 6. GESTIÓN DE USUARIOS (Solo Admin)
# ==========================================

@login_required
def gestion_usuarios(request):
    """Vista para gestionar usuarios (solo superusuarios)"""
    if not request.user.is_superuser:
        return render(
            request,
            "calificaciones/forbidden.html",
            {
                "title": "Acceso restringido",
                "reason": "Solo administradores pueden gestionar usuarios.",
                "action": "Contacta a tu administrador."
            },
            status=403,
        )
    
    from django.contrib.auth.models import User
    usuarios = User.objects.all().order_by('-date_joined')
    return render(request, 'calificaciones/gestion_usuarios.html', {'usuarios': usuarios})


@login_required
def crear_usuario(request):
    """Crea un usuario según el tipo especificado"""
    if not request.user.is_superuser:
        messages.error(request, "No tienes permisos para crear usuarios.")
        return redirect('dashboard')
    
    if request.method == "POST":
        from django.contrib.auth.models import User
        
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        tipo_usuario = request.POST.get('tipo_usuario', '')
        
        if not username or not password:
            messages.error(request, "Usuario y contraseña son obligatorios.")
            return redirect('gestion_usuarios')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, f"El usuario '{username}' ya existe.")
            return redirect('gestion_usuarios')
        
        # Determinar permisos según tipo
        is_staff = (tipo_usuario == 'analista')
        
        User.objects.create_user(
            username=username,
            password=password,
            first_name=first_name,
            last_name=last_name,
            is_staff=is_staff,
            is_superuser=False
        )
        
        rol = "Analista" if is_staff else "Corredor (Solo Lectura)"
        messages.success(request, f"Usuario '{username}' creado como {rol}.")
        return redirect('gestion_usuarios')
    
    return redirect('gestion_usuarios')


@login_required
def eliminar_usuario(request, user_id):
    """Elimina un usuario (no permite eliminar superusuarios)"""
    if not request.user.is_superuser:
        messages.error(request, "No tienes permisos para eliminar usuarios.")
        return redirect('dashboard')
    
    if request.method == "POST":
        from django.contrib.auth.models import User
        
        usuario = get_object_or_404(User, id=user_id)
        
        if usuario.is_superuser:
            messages.error(request, "No puedes eliminar usuarios administradores.")
            return redirect('gestion_usuarios')
        
        username = usuario.username
        usuario.delete()
        messages.success(request, f"Usuario '{username}' eliminado correctamente.")
    
    return redirect('gestion_usuarios')